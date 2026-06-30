# -*- coding: utf-8 -*-
"""openpyxl エンジン = Excel 無しでネイティブ Excel グラフ（OOXML）を生成する。

純 Python。Windows/macOS/Linux・CI/ヘッドレスで動く。生成されるグラフは
本物の編集可能な Excel グラフだが、書式の再現度には天井がある（COM/VBA が
できることの一部は OOXML に露出していない）。Excel が使えない提出環境向けの
フォールバック。COM エンジンと同じ ChartSpec を解釈する。

対応: 折れ線 / 散布図 / 棒 / 横棒 / 積み上げ棒 / 円 / ヒストグラム（事前ビン化）
制限: 第2軸は combine で対応。複合（系列ごとの種別混在）は近似的に扱う。
"""
import importlib.util
import os

import pandas as pd

from . import mapping as M

_EMU_PER_PT = 12700  # openpyxl の線幅は EMU 指定


def available():
    return importlib.util.find_spec("openpyxl") is not None


def write(spec, df, out_path):
    """ChartSpec + DataFrame からネイティブ Excel グラフ入り .xlsx を保存する。"""
    try:
        from openpyxl import Workbook
    except Exception as e:  # pragma: no cover
        raise RuntimeError("openpyxl が必要です（pip install openpyxl）。") from e

    out_path = os.path.abspath(out_path)
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    layout = _layout(spec)
    headers = [name for name, _c in layout]
    cols = [_column_values(df, col) for _name, col in layout]
    nrow = max((len(c) for c in cols), default=0)

    ws.append(headers)
    for i in range(nrow):
        ws.append([(c[i] if i < len(c) else None) for c in cols])

    chart = _build_chart(spec, df, ws, layout, nrow)
    if chart is not None:
        ws.add_chart(chart, _anchor(len(headers)))

    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- 内部
def _uses_x(spec):
    return spec.chart_type not in ("ヒストグラム", "箱ひげ")


def _layout(spec):
    layout = []
    if _uses_x(spec):
        layout.append((spec.xlabel or spec.x_col or "X", spec.x_col))
    for s in spec.series:
        layout.append((s.name, s.y_col))
    return layout


def _column_values(df, col):
    if col is None:
        return []
    if col not in df.columns:
        return []
    s = df[col]
    num = pd.to_numeric(s, errors="coerce")
    if num.notna().mean() >= 0.8:
        return [None if pd.isna(v) else float(v) for v in num.to_numpy()]
    return [None if pd.isna(v) else str(v) for v in s.to_numpy()]


def _anchor(ncols):
    """データ列の右隣にグラフを置くセルアンカー。"""
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(ncols + 2)}2"


def _x_is_numeric(df, spec):
    if not spec.x_col or spec.x_col not in df.columns:
        return True
    return bool(pd.to_numeric(df[spec.x_col], errors="coerce").notna().mean() >= 0.8)


def _col_pos(spec):
    """レイアウト上の 1 始まり列番号: (x_col_idx or None, {series: y_idx})。"""
    pos = 1
    x_idx = None
    if _uses_x(spec):
        x_idx = 1
        pos = 2
    ymap = {}
    for i, s in enumerate(spec.series):
        ymap[id(s)] = pos + i
    return x_idx, ymap


def _build_chart(spec, df, ws, layout, nrow):
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference

    ct = spec.chart_type
    x_idx, ymap = _col_pos(spec)
    maxr = 1 + nrow

    if ct == "円":
        chart = PieChart()
        s = spec.series[0]
        data = Reference(ws, min_col=ymap[id(s)], min_row=1, max_row=maxr)
        cats = Reference(ws, min_col=x_idx, min_row=2, max_row=maxr) if x_idx else None
        chart.add_data(data, titles_from_data=True)
        if cats is not None:
            chart.set_categories(cats)
        if spec.data_labels or spec.pct:
            from openpyxl.chart.label import DataLabelList
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = not spec.pct
            chart.dataLabels.showPercent = spec.pct
        _titles(chart, spec)
        _legend(chart, spec)
        return chart

    if ct in ("棒", "横棒", "積み上げ棒"):
        chart = BarChart()
        chart.type = "bar" if ct == "横棒" else "col"
        chart.grouping = "stacked" if ct == "積み上げ棒" else "clustered"
        if ct == "積み上げ棒":
            chart.overlap = 100
        for s in spec.series:
            data = Reference(ws, min_col=ymap[id(s)], min_row=1, max_row=maxr)
            chart.add_data(data, titles_from_data=True)
        if x_idx:
            cats = Reference(ws, min_col=x_idx, min_row=2, max_row=maxr)
            chart.set_categories(cats)
        _style_bar_series(chart, spec)
        _finish(chart, spec, df)
        return chart

    if ct == "ヒストグラム":
        return _build_histogram(spec, df, ws)

    # 折れ線 / 散布図（数値 X は散布図系で X を数値軸として保持）
    is_scatter = (ct == "散布図") or (ct == "折れ線" and _x_is_numeric(df, spec))
    if is_scatter:
        chart = _assemble(spec, df, x_idx=x_idx, ymap=ymap, ws=ws, nrow=nrow)
        _finish(chart, spec, df)
        return chart

    # カテゴリ折れ線
    chart = LineChart()
    for s in spec.series:
        data = Reference(ws, min_col=ymap[id(s)], min_row=1, max_row=maxr)
        chart.add_data(data, titles_from_data=True)
    if x_idx:
        cats = Reference(ws, min_col=x_idx, min_row=2, max_row=maxr)
        chart.set_categories(cats)
    for ser, s in zip(chart.series, spec.series):
        _style_xy_series(ser, s, spec, line=True)
    _finish(chart, spec, df)
    return chart


def _assemble(spec, df, *, x_idx, ymap, ws, nrow):
    """主軸/第2軸を分けて散布図を組み立て、必要なら combine する。

    ScatterChart は X/Y とも値軸なので、第2軸を出すには第2チャートに
    「隠しX軸(新 axId, delete=True)＋新Y軸(crosses='max')」を与える必要がある
    （実機 Excel 検証でこの組み合わせのみ系列が第2軸 group=2 になった）。
    """
    from openpyxl.chart import ScatterChart, Reference, Series
    maxr = 1 + nrow

    xref = Reference(ws, min_col=x_idx, min_row=2, max_row=maxr) if x_idx \
        else Reference(ws, min_col=ymap[id(spec.series[0])], min_row=2, max_row=maxr)

    prim = ScatterChart()
    prim.x_axis.delete = False
    prim.y_axis.delete = False
    sec = None
    for s in spec.series:
        yref = Reference(ws, min_col=ymap[id(s)], min_row=1, max_row=maxr)
        ser = Series(yref, xref, title_from_data=True)
        _style_xy_series(ser, s, spec,
                         line=(spec.chart_type == "折れ線" and s.kind != "scatter"))
        if s.axis == "secondary":
            if sec is None:
                sec = ScatterChart()
            sec.series.append(ser)
        else:
            prim.series.append(ser)

    # 全系列が第2軸指定などで主軸が空なら、主軸へ昇格させて単一チャートに
    if not prim.series and sec is not None:
        prim, sec = sec, None
        prim.x_axis.delete = False
        prim.y_axis.delete = False

    if sec is not None and sec.series:
        sec.x_axis.axId = 500       # 第2の（隠し）X軸
        sec.x_axis.delete = True
        sec.y_axis.axId = 200       # 第2のY軸
        sec.y_axis.delete = False
        sec.y_axis.crosses = "max"  # 右側に表示
        if spec.secondary_label:
            sec.y_axis.title = spec.secondary_label
        prim += sec
    return prim


def _build_histogram(spec, df, ws):
    """ヒストグラム: 値列をビン化して棒グラフとして出力（事前集計）。"""
    import numpy as np
    from openpyxl.chart import BarChart, Reference

    # 別シートにビン集計を書く
    hs = ws.parent.create_sheet("Hist")
    bins = int(spec.bins or 10)
    # 全系列の値域でビン境界を共有
    allvals = []
    for s in spec.series:
        v = pd.to_numeric(df[s.y_col], errors="coerce").dropna().to_numpy()
        allvals.append(v)
    flat = np.concatenate(allvals) if allvals else np.array([0.0, 1.0])
    edges = np.histogram_bin_edges(flat, bins=bins)
    centers = (edges[:-1] + edges[1:]) / 2

    hs.cell(1, 1, "bin")
    for i, c in enumerate(centers, start=2):
        hs.cell(i, 1, float(round(c, 6)))
    for j, s in enumerate(spec.series, start=2):
        hs.cell(1, j, s.name)
        v = pd.to_numeric(df[s.y_col], errors="coerce").dropna().to_numpy()
        counts, _ = np.histogram(v, bins=edges)
        for i, cnt in enumerate(counts, start=2):
            hs.cell(i, j, int(cnt))

    chart = BarChart()
    chart.type = "col"
    chart.gapWidth = 20
    maxr = 1 + len(centers)
    data = Reference(hs, min_col=2, max_col=1 + len(spec.series), min_row=1, max_row=maxr)
    cats = Reference(hs, min_col=1, min_row=2, max_row=maxr)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _style_bar_series(chart, spec)
    chart.title = spec.title or None
    if spec.xlabel:
        chart.x_axis.title = spec.xlabel
    chart.y_axis.title = spec.ylabel or "頻度"
    _legend(chart, spec)
    return chart


# --- 系列スタイル -----------------------------------------------------------
def _style_xy_series(ser, s, spec, *, line):
    from openpyxl.chart.marker import Marker
    from openpyxl.drawing.line import LineProperties

    rgb = M.hex_to_rrggbb(s.color)
    want_line = line and M.openpyxl_dash(s.linestyle) is not None
    if want_line:
        gp = ser.graphicalProperties
        if rgb:
            gp.line.solidFill = rgb
        gp.line.width = int(max(0.25, s.linewidth) * _EMU_PER_PT)
        dash = M.openpyxl_dash(s.linestyle)
        if dash and dash != "solid":
            gp.line.prstDash = dash
        ser.smooth = False
    else:
        # 線なし（散布図・線種なし）
        ser.graphicalProperties.line = LineProperties(noFill=True)

    # マーカー
    sym = M.openpyxl_marker(s.marker)
    if spec.chart_type == "散布図" and (not s.marker):
        sym = "circle"
    if sym and sym != "none":
        ser.marker = Marker(symbol=sym, size=int(max(2, min(72, round(s.markersize)))))
        if rgb:
            ser.marker.graphicalProperties.solidFill = rgb
            ser.marker.graphicalProperties.line.solidFill = rgb

    if spec.data_labels:
        from openpyxl.chart.label import DataLabelList
        ser.dLbls = DataLabelList()
        ser.dLbls.showVal = True


def _style_bar_series(chart, spec):
    from openpyxl.chart.shapes import GraphicalProperties

    for ser, s in zip(chart.series, spec.series):
        rgb = M.hex_to_rrggbb(s.color)
        if rgb:
            ser.graphicalProperties = GraphicalProperties(solidFill=rgb)
    if spec.data_labels:
        from openpyxl.chart.label import DataLabelList
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True


# --- 体裁（タイトル・軸・凡例）共通 -----------------------------------------
def _titles(chart, spec):
    if spec.title:
        chart.title = spec.title


def _legend(chart, spec):
    if not spec.legend:
        chart.legend = None
    elif chart.legend is not None:
        chart.legend.position = M.openpyxl_legend_position(spec.legend_loc)


def _finish(chart, spec, df):
    """折れ線/散布/棒の軸・タイトル・凡例を仕上げる。"""
    if spec.title:
        chart.title = spec.title
    try:
        if spec.xlabel:
            chart.x_axis.title = spec.xlabel
        if spec.ylabel:
            chart.y_axis.title = spec.ylabel
        # 軸の表示を確実化（openpyxl は delete=True だと軸ラベルが出ないことがある）
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        if spec.xmin is not None:
            chart.x_axis.scaling.min = spec.xmin
        if spec.xmax is not None:
            chart.x_axis.scaling.max = spec.xmax
        if spec.ymin is not None:
            chart.y_axis.scaling.min = spec.ymin
        if spec.ymax is not None:
            chart.y_axis.scaling.max = spec.ymax
        if spec.xlog:
            chart.x_axis.scaling.logBase = 10
        if spec.ylog:
            chart.y_axis.scaling.logBase = 10
        if spec.xinvert:
            chart.x_axis.scaling.orientation = "maxMin"
        if spec.yinvert:
            chart.y_axis.scaling.orientation = "maxMin"
        if spec.grid:
            from openpyxl.chart.axis import ChartLines
            chart.y_axis.majorGridlines = ChartLines()
    except Exception:
        pass
    _legend(chart, spec)
